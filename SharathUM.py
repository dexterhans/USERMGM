import openpyxl as openpyxl
from pymongo import MongoClient

stageConnectionStr = "mongodb://mms-automation:GNYbrmR99XkJHPtAEmxSFyCr@superprodmongo-0-pvt.infra.joveo.com:27017,superprodmongo-4-pvt.infra.joveo.com:27017,superprodmongo-5-pvt.infra.joveo.com:27017/admin?readPreference=secondary&connectTimeoutMS=10000&authSource=admin&authMechanism=SCRAM-SHA-1"

client2 = MongoClient(stageConnectionStr)

collection = client2.mojo.joveo_users

wb = openpyxl.load_workbook("/Users/deepakranganathan/Downloads/User Migration - User Management-6.xlsx")


sheets = wb.sheetnames
sheet = wb[sheets[13]]
print(sheet)

for row in range(4, sheet.max_row):
    if (sheet.cell(row, 2).value != None):
        collection.update_one({"email": sheet.cell(2, 2).value},
                              {"$push": {"scope": {"id": sheet.cell(row, 2).value, "application": "Mojo",
                                                   "clients": [], "publishers": []}}})
    else:
        break