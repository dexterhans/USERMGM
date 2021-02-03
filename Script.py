import getopt
import sys
import openpyxl as openpyxl
from pymongo import MongoClient
import shelve


# providing commandline arguments to input connection string and excel sheet
def main(argv):
    global connectionStr
    global excelSheet
    try:
        opts, args = getopt.getopt(argv, "hc:s:", ["connectionStr=", "sheet="])
    except getopt.GetoptError:
        print('test.py -c <connectionStr> -s <sheet>')
        sys.exit(2)
    for opt, arg in opts:
        if opt == '-h':
            print('test.py -c <connectionStr> -s <sheet>')
            sys.exit()
        elif opt in ("-c", "--connectionStr"):
            connectionStr = arg
        elif opt in ("-s", "--sheet"):
            excelSheet = arg
    print('connectionStr is ', connectionStr)
    print('excelSheet is ', excelSheet)


connectionStr = ''
excelSheet = ''

if __name__ == "__main__":
    main(sys.argv[1:])

# stageConnectionStr = "mongodb://mms-automation:GNYbrmR99XkJHPtAEmxSFyCr@services-mongo.infra-dev.joveo.com:27000/admin?connectTimeoutMS=10000&authSource=admin&authMechanism=SCRAM-SHA-1"

client2 = MongoClient(connectionStr)

# collection to work on
collection = client2.mojo.joveo_users

d = shelve.open("persistent")

if d.get("backup", 0) == 0:
    d['backup'] = 1
    collection.aggregate([{'$match': {}}, {'$out': "joveo_users_UM_backup"}])
# sheet = "/Users/deepakranganathan/Downloads/User Migration - User Management-2.xlsx"

wb = openpyxl.load_workbook(excelSheet)

# 1st usecase
# remove mojo application form scope of the rows marked in grey,
# and if the rows do not contain any other roles delete it
sheets = wb.sheetnames
sheet = wb[sheets[1]]

greyEmail = set()
for row in range(2, sheet.max_row):
    if sheet.cell(row, 4).fill.bgColor.rgb == "FFCCCCCC":
        greyEmail.add(sheet.cell(row, 4).value)

for val in greyEmail:
    collection.update_one({"email": val},
                          {"$pull": {"scope": {"application": "Mojo"}}})
    print(val)

# 2nd usecase
# add agencies to the user email mentioned in the sheet
sheet = wb[sheets[5]]

for col in range(2, sheet.max_column):
    if (collection.find_one({"email": sheet.cell(2, col).value}) != None):
        for row in range(4, sheet.max_row):
            if (sheet.cell(row, col).value != None):
                collection.update_one({"email": sheet.cell(2, col).value},
                                      {"$push": {"scope": {"id": sheet.cell(row, col).value, "application": "Mojo",
                                                           "clients": [], "publishers": []}}})
                collection.update_one({"email": sheet.cell(2, col).value}, {'$set': {
                    "roles.0": sheet.cell(3, col).value}})
            else:
                break
    else:
        print("email {} is not available".format(sheet.cell(2, col).value))

# print("deleting:")
for val in greyEmail:
    collection.delete_one({"$and": [{"scope": {"$size": 0}},
                                    {"email": val}]})

# 3rd usecase
# obtaining old role and new role as key, value pair in a dict
dict = {}
sheetx = wb[sheets[9]]
for row in range(2, sheetx.max_row):
    if (sheetx.cell(row, 1).value != None):
        dict.update({sheetx.cell(row, 1).value: sheetx.cell(row, 2).value})
        print(sheetx.cell(row, 1).value)
    else:
        break

print()
print("updating roles")
# updating the roles
sheetx = wb[sheets[1]]
print(sheetx)
for row in range(2, sheetx.max_row):
    if (sheetx.cell(row, 4).value != None):
        print(sheetx.cell(row, 4).value)
        print(sheetx.cell(row, 4).fill.bgColor.rgb)
        cursor = collection.find_one({"email": sheetx.cell(row, 4).value})
        if sheetx.cell(row, 4).fill.bgColor.rgb == "00000000" and cursor is not None and cursor['roles'] != []:
            print(sheetx.cell(row, 4).value)
            print(cursor['roles'])
            print(cursor['roles'][0])
            print("email is {} and roles is {}".format(sheetx.cell(row, 4).value, cursor['roles'][0]))
            print()
            print()
            collection.update_one({"email": sheetx.cell(row, 4).value},
                                  {'$set': {
                                      "roles.0": dict.get(cursor['roles'][0], cursor['roles'][0])
                                  }})
    else:
        break
