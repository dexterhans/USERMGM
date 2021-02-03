import shelve
import datetime
import sys
import uuid

from pymongo import MongoClient

import openpyxl as openpyxl

stageConnectionStr = "mongodb://mms-automation:GNYbrmR99XkJHPtAEmxSFyCr@superprodmongo-0-pvt.infra.joveo.com:27017,superprodmongo-4-pvt.infra.joveo.com:27017,superprodmongo-5-pvt.infra.joveo.com:27017/admin?readPreference=secondary&connectTimeoutMS=10000&authSource=admin&authMechanism=SCRAM-SHA-1"

client2 = MongoClient(stageConnectionStr)

# collection to work on
collection = client2.mojo.roles

d = shelve.open("shelve_persistent")

if d.get("backup", 0) == 0:
    d['backup'] = 1
    collection.aggregate([{'$match': {}}, {'$out': "roles_backup"}])

wb = openpyxl.load_workbook("/Users/deepakranganathan/Downloads/User Migration - User Management-4.xlsx")

sheets = wb.sheetnames
sheet = wb[sheets[8]]

for row in range(3,16):
    list = []
    for col in range(5,18):
        if sheet.cell(row,col).value != None:
            list.append(sheet.cell(row,col).value)
        else:
            break
    try:
        collection.insert_one({
        "_id": str(uuid.uuid1()),
        "roleName":sheet.cell(row,1).value,
        "description":sheet.cell(row,2).value,
        "key":sheet.cell(row,3).value,
        "createdBy":"deepak",
        "createdOn":datetime.datetime.utcnow(),
        "isActive":True,
        "permissions":list,
        "roleFamily":sheet.cell(row,18).value
        })
    except:
        print("Oops!", sys.exc_info()[0], "occurred.")
